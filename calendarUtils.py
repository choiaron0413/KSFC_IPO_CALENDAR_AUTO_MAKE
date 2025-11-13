from openpyxl import styles 
from openpyxl.styles.borders import Border, Side 
from openpyxl.styles import Font, Color, Alignment 
from openpyxl import Workbook 
import openpyxl 
import calendar 
import datetime 


BORDER_NONE = None 
BORDER_DASHDOT = 'dashDot' 
BORDER_DASHDOTDOT = 'dashDotDot' 
BORDER_DASHED = 'dashed' 
BORDER_DOTTED = 'dotted' 
BORDER_DOUBLE = 'double' 
BORDER_HAIR = 'hair' 
BORDER_MEDIUM = 'medium' 
BORDER_MEDIUMDASHDOT = 'mediumDashDot' 
BORDER_MEDIUMDASHDOTDOT = 'mediumDashDotDot' 
BORDER_MEDIUMDASHED = 'mediumDashed' 
BORDER_SLANTDASHDOT = 'slantDashDot' 
BORDER_THICK = 'thick' 
BORDER_THIN = 'thin' 

    # set font style for each cell
def cell_font_style(font_color, font_size, align, font_bold_yn = False) : 
    font = Font(color = font_color, size = font_size, bold = font_bold_yn ) 
    alignment = Alignment(horizontal="center", vertical= align) 
    return font, alignment 


    # set_titile 
def set_title(ws, title) : 
    style = cell_font_style("000000", 20, "center", True) 
    ws["A1"] = title 
    A1 = ws["A1"] 
    A1.font = style[0] 
    A1.alignment = style[1] 
    
    ws.row_dimensions[1].height = 50 
    ws.merge_cells("A1:G1") 
    
 
    # set_weeks 
def set_weeks_title(ws) : 
    weeks = ["일", "월","화","수", "목", "금", "토"] 
     
    for w in range(1,8) : 
        #c = ws.cell(row=3, column=w, value=weeks[w-1]) 
        c = ws.cell(row=2, column=w, value=weeks[w-1]) 
        
        if 1 == w : 
            style = cell_font_style("FF0000", 14, "center", True) 
        elif 7 == w :
            style = cell_font_style("0000FF", 14, "center", True) 
        else :
            style = cell_font_style("000000", 14, "center", True) 
        
        c.font = style[0] 
        c.alignment = style[1] 
        c.border = Border( 
            left=Side(style=BORDER_NONE)
            , right=Side(style=BORDER_NONE)
            , top=Side(style=BORDER_THIN)
            , bottom=Side(style=BORDER_THIN)
        );
        
    ws.row_dimensions[3].height = 37.5

    
    #get_border 
def get_border(top, right, bottom, left) : 
    thin_border = Border( 
        left=Side(style=left)
        , right=Side(style=right)
        , top=Side(style=top)
        , bottom=Side(style=bottom)
    )
    
    return thin_border 
    
    
    #set_calendar
def set_ipo_calendar(ws, year, month, ipo_df) : 
    o = calendar.Calendar(calendar.SUNDAY) 
    
    #i=4 # current excel row
    i=3 # current excel row
    
    for weeks in o.monthdays2calendar(year, month) : 
        j=1 # current excel column ==> 일, 월, 화 .... , 토
        
        sbsc_ext_yn = False # 해당일자 청약정보 존재여부
        max_height_yn = False # 하루에 중복청약 대상 종목이 2개이상인 경우 해당 주의 높이를 더 크게 하기 위함
        print("AAAAA")
        for week in weeks : 
            
            if ( week[0] > 0 ) : 
                #ws.cell(row=i, column=j, value=week[0]) 
                print("BBBBB")
                temp_sbsc_nd_dt = str(year) + "-" + str(month).zfill(2) + "-" + str(week[0]).zfill(2)
                temp_ipo_df=ipo_df.loc[ ipo_df['청약종료일자'].str.contains(temp_sbsc_nd_dt), : ]
                temp_value = str(week[0])
                print("CCCCC")
                if len( temp_ipo_df ) > 0 :
                    
                    sbsc_ext_yn = True
                    
                    if len( temp_ipo_df ) > 1 :
                        max_height_yn = True
                    idx_sbsc_cnt_per_date = 1
                    
                    for itr, row in temp_ipo_df.iterrows() :
                        temp_value = temp_value + "\n" 
                        
                        if idx_sbsc_cnt_per_date != 1 :
                            temp_value = temp_value + "\n" # 종목이 바뀔 때 개행문자
                        
                        temp_value = temp_value + '- ' + row['종목명']
                        
                        temp_writer_list=row['주간사'].split(",")
                        #print(temp_writer_list)
                        
                        # 종목별 주간사내역 입력
                        for temp_writer in temp_writer_list :
                            temp_value = temp_value + "\n" 
                            temp_value = temp_value + "  * " + temp_writer
                            
                        idx_sbsc_cnt_per_date = idx_sbsc_cnt_per_date + 1
                           
                    print( 
                       str(year) + # yyyy-
                       "-" + 
                       str(month).zfill(2) + #mm- 
                       "-" #+ 
                       #str(week[0]).zfill(2) # dd
                    )
                    print(temp_value)
                    
                    temp_cell=ws.cell(row=i, column=j, value=temp_value) 
                    
                    temp_cell.font = Font( size = 11, bold=True ) 
                    temp_cell.alignment = Alignment(horizontal="left", vertical= "top", wrapText=True) 
                    
                    
                else :
                    temp_cell=ws.cell(row=i, column=j, value=week[0]) 
                    temp_cell.font = Font( size = 10 ) 
                    temp_cell.alignment = Alignment(horizontal="left", vertical= "top", wrapText=True) 
                
                temp_cell.border = get_border(BORDER_DOTTED, BORDER_DOTTED, BORDER_DOTTED, BORDER_DOTTED) 
            
            if 1 == j :
                ws.column_dimensions['A'].width = 25
            elif 2== j :
                ws.column_dimensions['B'].width = 25
            elif 3== j :
                ws.column_dimensions['C'].width = 25
            elif 4== j :
                ws.column_dimensions['D'].width = 25
            elif 5== j :
                ws.column_dimensions['E'].width = 25
            elif 6== j :
                ws.column_dimensions['F'].width = 25
            else :
                ws.column_dimensions['G'].width = 25
            
            j = j + 1
        print("DDDDD")
        if True == sbsc_ext_yn :
            if True == max_height_yn :
                ws.row_dimensions[i].height = 210
            else :
                ws.row_dimensions[i].height = 170
        else :
            ws.row_dimensions[i].height = 50
            
        i = i + 1  # current excel row = current excel row + 1
                   # current excel row ==> '1주', '2주' ....